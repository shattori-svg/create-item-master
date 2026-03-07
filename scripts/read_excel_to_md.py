# -*- coding: utf-8 -*-
"""
Excelファイルの内容をMarkdown形式で出力する。
要件定義書から参照しやすいように、プロジェクト直下に .md を生成する。
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl をインストールしてください: pip install openpyxl")
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parent.parent


def sheet_to_markdown(ws, max_rows=100, max_cols=50):
    """ワークシートをMarkdownの表形式に変換"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return "（データなし）\n"
    # 空でない列までに限定
    col_count = max(len(r) for r in rows) if rows else 0
    col_count = min(col_count, max_cols)
    out = []
    for i, row in enumerate(rows[:max_rows]):
        if row is None:
            continue
        cells = list(row)[:col_count]
        cells = ["" if c is None else str(c).strip() for c in cells]
        out.append("| " + " | ".join(cells) + " |")
        if i == 0:
            out.append("| " + " | ".join("---" for _ in cells) + " |")
    return "\n".join(out) + "\n"


def dump_workbook(path: Path, out_path: Path, max_rows_per_sheet=80):
    """ブック全体をダンプ"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    lines = [f"# {path.name}\n", f"元ファイル: `{path.name}`\n"]
    for name in wb.sheetnames:
        ws = wb[name]
        lines.append(f"\n## シート: {name}\n")
        lines.append(sheet_to_markdown(ws, max_rows=max_rows_per_sheet))
    wb.close()
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"出力: {out_path}")


def main():
    spec_file = PROJECT_ROOT / "OIC Tech Spec Item Import Template Specification 20260116.xlsx"
    sample_file = PROJECT_ROOT / "Item_260306 0906 (1).xlsx"

    if spec_file.exists():
        dump_workbook(
            spec_file,
            PROJECT_ROOT / "docs" / "OIC_Tech_Spec_Item_Import_Template_Specification_20260116.md",
            max_rows_per_sheet=120,
        )
    else:
        print(f"仕様書が見つかりません: {spec_file}")

    if sample_file.exists():
        dump_workbook(
            sample_file,
            PROJECT_ROOT / "docs" / "Item_260306_0906_sample.md",
            max_rows_per_sheet=150,
        )
    else:
        print(f"サンプルが見つかりません: {sample_file}")


if __name__ == "__main__":
    (PROJECT_ROOT / "docs").mkdir(exist_ok=True)
    main()
